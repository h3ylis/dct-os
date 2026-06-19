import csv
import hashlib
import io
import os

from flask import Blueprint, Response, jsonify, request, send_file

from dct_os.db import get_db, register_supplier

bp = Blueprint("dockets", __name__)


def _build_docket(db, header_id):
    header = db.execute(
        """SELECT dh.*,
                  po.number AS po_number
           FROM docket_headers dh
           LEFT JOIN purchase_orders po ON dh.purchase_order_id = po.id
           WHERE dh.id = ?""",
        (header_id,),
    ).fetchone()
    if header is None:
        return None
    d = dict(header)
    lines = db.execute(
        """SELECT dl.*,
                  wo.number AS wo_number,
                  wo.description AS wo_description,
                  cc.code AS cost_code,
                  cc.description AS cc_description,
                  r.description AS resource_description
           FROM docket_lines dl
           LEFT JOIN work_orders wo ON dl.work_order_id = wo.id
           LEFT JOIN cost_codes cc ON dl.cost_code_id = cc.id
           LEFT JOIN resources r ON dl.resource_id = r.id
           WHERE dl.docket_id = ?
           ORDER BY dl.sort_order, dl.id""",
        (header_id,),
    ).fetchall()
    d["lines"] = [dict(ln) for ln in lines]
    d["line_count"] = len(d["lines"])
    d["total_amount"] = sum(ln["amount"] or 0 for ln in d["lines"])
    return d


@bp.route("/projects/<int:project_id>/dockets", methods=["GET"])
def list_dockets(project_id):
    db = get_db()
    headers = db.execute(
        """SELECT dh.*,
                  po.number AS po_number,
                  COUNT(dl.id) AS line_count,
                  COALESCE(SUM(dl.amount), 0) AS total_amount,
                  GROUP_CONCAT(DISTINCT wo.number) AS wo_numbers,
                  GROUP_CONCAT(DISTINCT cc.code) AS cost_codes
           FROM docket_headers dh
           LEFT JOIN purchase_orders po ON dh.purchase_order_id = po.id
           LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
           LEFT JOIN work_orders wo ON dl.work_order_id = wo.id
           LEFT JOIN cost_codes cc ON dl.cost_code_id = cc.id
           WHERE dh.project_id = ?
           GROUP BY dh.id
           ORDER BY dh.date DESC, dh.id DESC""",
        (project_id,),
    ).fetchall()
    result = []
    for h in headers:
        d = dict(h)
        lines = db.execute(
            """SELECT dl.*,
                      wo.number AS wo_number,
                      cc.code AS cost_code,
                      r.description AS resource_description
               FROM docket_lines dl
               LEFT JOIN work_orders wo ON dl.work_order_id = wo.id
               LEFT JOIN cost_codes cc ON dl.cost_code_id = cc.id
               LEFT JOIN resources r ON dl.resource_id = r.id
               WHERE dl.docket_id = ?
               ORDER BY dl.sort_order, dl.id""",
            (d["id"],),
        ).fetchall()
        d["lines"] = [dict(ln) for ln in lines]
        result.append(d)
    return jsonify(result)


@bp.route("/dockets/<int:docket_id>", methods=["GET"])
def get_docket(docket_id):
    db = get_db()
    d = _build_docket(db, docket_id)
    if d is None:
        return jsonify({"error": "Docket not found"}), 404
    return jsonify(d)


@bp.route("/projects/<int:project_id>/dockets", methods=["POST"])
def create_docket(project_id):
    db = get_db()
    data = request.get_json()
    if not data or not data.get("date"):
        return jsonify({"error": "date is required"}), 400

    cur = db.execute(
        """INSERT INTO docket_headers
           (project_id, purchase_order_id, supplier_name, date,
            docket_number, notes, source_hash, source_filename,
            source_filepath)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            project_id,
            data.get("purchase_order_id"),
            register_supplier(db, data.get("supplier_name")),
            data["date"],
            data.get("docket_number"),
            data.get("notes"),
            data.get("source_hash"),
            data.get("source_filename"),
            data.get("source_filepath"),
        ),
    )
    header_id = cur.lastrowid

    for i, line in enumerate(data.get("lines", [])):
        qty = line.get("qty", 0)
        rate = line.get("rate", 0)
        amount = line.get("amount") or qty * rate
        db.execute(
            """INSERT INTO docket_lines
               (docket_id, work_order_id, cost_code_id, resource_id,
                description, qty, unit, rate, amount, sort_order)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                header_id,
                line.get("work_order_id"),
                line.get("cost_code_id"),
                line.get("resource_id"),
                line.get("description"),
                qty,
                line.get("unit"),
                rate,
                amount,
                line.get("sort_order", i),
            ),
        )
    db.commit()
    return jsonify(_build_docket(db, header_id)), 201


@bp.route("/dockets/<int:docket_id>", methods=["PUT"])
def update_docket(docket_id):
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
    existing = db.execute(
        "SELECT * FROM docket_headers WHERE id = ?", (docket_id,)
    ).fetchone()
    if existing is None:
        return jsonify({"error": "Docket not found"}), 404

    header_fields = [
        "purchase_order_id", "supplier_name", "date",
        "docket_number", "notes", "source_hash", "source_filename",
        "source_filepath",
    ]
    updates = {f: data[f] for f in header_fields if f in data}
    if "supplier_name" in updates:
        updates["supplier_name"] = register_supplier(db, updates["supplier_name"])
    if updates:
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        set_clause += ", updated_at = datetime('now')"
        values = list(updates.values())
        values.append(docket_id)
        db.execute(
            f"UPDATE docket_headers SET {set_clause} WHERE id = ?", values
        )

    if "lines" in data:
        db.execute("DELETE FROM docket_lines WHERE docket_id = ?", (docket_id,))
        for i, line in enumerate(data["lines"]):
            qty = line.get("qty", 0)
            rate = line.get("rate", 0)
            amount = line.get("amount") or qty * rate
            db.execute(
                """INSERT INTO docket_lines
                   (docket_id, work_order_id, cost_code_id, resource_id,
                    description, qty, unit, rate, amount, sort_order)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    docket_id,
                    line.get("work_order_id"),
                    line.get("cost_code_id"),
                    line.get("resource_id"),
                    line.get("description"),
                    qty,
                    line.get("unit"),
                    rate,
                    amount,
                    line.get("sort_order", i),
                ),
            )
    db.commit()
    return jsonify(_build_docket(db, docket_id))


@bp.route("/dockets/<int:docket_id>", methods=["DELETE"])
def delete_docket(docket_id):
    db = get_db()
    existing = db.execute(
        "SELECT * FROM docket_headers WHERE id = ?", (docket_id,)
    ).fetchone()
    if existing is None:
        return jsonify({"error": "Docket not found"}), 404
    db.execute("DELETE FROM docket_headers WHERE id = ?", (docket_id,))
    db.commit()
    return jsonify({"deleted": docket_id})


@bp.route("/projects/<int:project_id>/summary", methods=["GET"])
def project_summary(project_id):
    db = get_db()
    summary = db.execute(
        """SELECT
               COUNT(DISTINCT dh.id) AS total_dockets,
               COALESCE(SUM(dl.amount), 0) AS total_spend,
               COUNT(DISTINCT dh.supplier_name) AS supplier_count,
               COUNT(DISTINCT dh.date) AS active_days
           FROM docket_headers dh
           LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
           WHERE dh.project_id = ?""",
        (project_id,),
    ).fetchone()
    return jsonify(dict(summary))


@bp.route("/projects/<int:project_id>/cost-report", methods=["GET"])
def cost_report(project_id):
    db = get_db()
    rows = db.execute(
        """SELECT cc.id, cc.code, cc.description, cc.budget_amount,
                  COALESCE(SUM(dl.amount), 0) AS actual_spend,
                  cc.budget_amount - COALESCE(SUM(dl.amount), 0) AS variance
           FROM cost_codes cc
           LEFT JOIN docket_lines dl ON dl.cost_code_id = cc.id
           LEFT JOIN docket_headers dh ON dl.docket_id = dh.id
                AND dh.project_id = cc.project_id
           WHERE cc.project_id = ?
           GROUP BY cc.id
           ORDER BY cc.code""",
        (project_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/projects/<int:project_id>/dashboard", methods=["GET"])
def project_dashboard(project_id):
    """One combined aggregation for the whole dashboard.

    Everything here is a rollup of *this project's own* docket data — no
    cross-project or rate-history aggregation (that's the commercial edition's
    edge). Returned as a single payload so the front-end renders every panel
    from one fetch:

        tiles        — headline budget / spent / remaining / dockets / suppliers
        cost_codes   — per-cost-code budget vs actual (the burn-down)
        spend_series — cumulative weekly spend (the time axis)
        wo_costs     — spend grouped by work order (the donut)
        po_drawdown  — per active PO, drawn vs committed
        top_suppliers— top 5 suppliers by spend
        claimed/to_claim — spend split by whether the docket is claimed
    """
    db = get_db()

    # --- Cost codes (budget vs actual) — same query as /cost-report ---
    cost_codes = [
        dict(r)
        for r in db.execute(
            """SELECT cc.id, cc.code, cc.description, cc.budget_amount,
                      COALESCE(SUM(dl.amount), 0) AS actual_spend
               FROM cost_codes cc
               LEFT JOIN docket_lines dl ON dl.cost_code_id = cc.id
               LEFT JOIN docket_headers dh ON dl.docket_id = dh.id
                    AND dh.project_id = cc.project_id
               WHERE cc.project_id = ?
               GROUP BY cc.id
               ORDER BY cc.code""",
            (project_id,),
        ).fetchall()
    ]

    # --- Tiles — headline totals (budget from cost codes, rest from dockets) ---
    summary = db.execute(
        """SELECT COUNT(DISTINCT dh.id) AS total_dockets,
                  COALESCE(SUM(dl.amount), 0) AS total_spend,
                  COUNT(DISTINCT dh.supplier_name) AS supplier_count
           FROM docket_headers dh
           LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
           WHERE dh.project_id = ?""",
        (project_id,),
    ).fetchone()
    total_budget = sum(c["budget_amount"] or 0 for c in cost_codes)
    total_spent = summary["total_spend"] or 0
    tiles = {
        "total_budget": total_budget,
        "total_spent": total_spent,
        "remaining": total_budget - total_spent,
        "total_dockets": summary["total_dockets"] or 0,
        "supplier_count": summary["supplier_count"] or 0,
    }

    # --- Spend over time — cumulative weekly burn-up ---
    week_rows = db.execute(
        """SELECT strftime('%Y-%W', dh.date) AS week,
                  MIN(dh.date) AS week_start,
                  COALESCE(SUM(dl.amount), 0) AS amount
           FROM docket_headers dh
           JOIN docket_lines dl ON dl.docket_id = dh.id
           WHERE dh.project_id = ? AND dh.date IS NOT NULL AND dh.date != ''
           GROUP BY week
           ORDER BY week""",
        (project_id,),
    ).fetchall()
    spend_series = []
    cumulative = 0
    for r in week_rows:
        cumulative += r["amount"] or 0
        spend_series.append({
            "week": r["week"],
            "week_start": r["week_start"],
            "amount": r["amount"] or 0,
            "cumulative": cumulative,
        })

    # --- Cost by work order (the donut) ---
    wo_costs = [
        dict(r)
        for r in db.execute(
            """SELECT wo.id, wo.number, wo.description,
                      COALESCE(SUM(dl.amount), 0) AS amount
               FROM work_orders wo
               JOIN docket_lines dl ON dl.work_order_id = wo.id
               JOIN docket_headers dh ON dl.docket_id = dh.id
                    AND dh.project_id = wo.project_id
               WHERE wo.project_id = ?
               GROUP BY wo.id
               HAVING amount > 0
               ORDER BY amount DESC""",
            (project_id,),
        ).fetchall()
    ]

    # --- PO drawdown — per active PO, drawn vs committed ---
    po_drawdown = [
        dict(r)
        for r in db.execute(
            """SELECT po.id, po.number, po.supplier_name,
                      po.value AS committed,
                      COALESCE(SUM(dl.amount), 0) AS drawn
               FROM purchase_orders po
               LEFT JOIN docket_headers dh ON dh.purchase_order_id = po.id
               LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
               WHERE po.project_id = ? AND po.is_active = 1
               GROUP BY po.id
               ORDER BY po.number""",
            (project_id,),
        ).fetchall()
    ]

    # --- Top suppliers by spend ---
    top_suppliers = [
        dict(r)
        for r in db.execute(
            """SELECT dh.supplier_name AS name,
                      COALESCE(SUM(dl.amount), 0) AS amount
               FROM docket_headers dh
               JOIN docket_lines dl ON dl.docket_id = dh.id
               WHERE dh.project_id = ?
                 AND dh.supplier_name IS NOT NULL AND dh.supplier_name != ''
               GROUP BY dh.supplier_name
               ORDER BY amount DESC
               LIMIT 5""",
            (project_id,),
        ).fetchall()
    ]

    # --- Claimed vs to-claim — cash-flow split off the docket claim tags ---
    claim_row = db.execute(
        """SELECT
               COALESCE(SUM(CASE WHEN dh.claimed_reference IS NOT NULL
                                  AND dh.claimed_reference != ''
                                 THEN dl.amount ELSE 0 END), 0) AS claimed,
               COALESCE(SUM(CASE WHEN dh.claimed_reference IS NULL
                                  OR dh.claimed_reference = ''
                                 THEN dl.amount ELSE 0 END), 0) AS to_claim
           FROM docket_headers dh
           LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
           WHERE dh.project_id = ?""",
        (project_id,),
    ).fetchone()

    return jsonify({
        "tiles": tiles,
        "cost_codes": cost_codes,
        "spend_series": spend_series,
        "wo_costs": wo_costs,
        "po_drawdown": po_drawdown,
        "top_suppliers": top_suppliers,
        "claimed": claim_row["claimed"],
        "to_claim": claim_row["to_claim"],
    })


@bp.route("/projects/<int:project_id>/suppliers", methods=["GET"])
def list_project_suppliers(project_id):
    """The canonical supplier list — the suppliers reference table is the
    single source of truth shared across every screen and project.
    (project_id is kept in the route for client compatibility.)
    """
    db = get_db()
    rows = db.execute(
        "SELECT name FROM suppliers ORDER BY name COLLATE NOCASE"
    ).fetchall()
    return jsonify([r["name"] for r in rows])


def _build_summary_filter(project_id, args):
    """Build WHERE clause for summary queries (shared by JSON + CSV)."""
    supplier = args.get("supplier")
    if not supplier:
        return None, None, "supplier parameter is required"

    date_from = args.get("date_from")
    date_to = args.get("date_to")
    docket_ids_raw = args.get("docket_ids")

    params = [project_id, supplier]
    where = "dh.project_id = ? AND dh.supplier_name = ?"

    if docket_ids_raw:
        ids = [int(x) for x in docket_ids_raw.split(",") if x.strip().isdigit()]
        if ids:
            placeholders = ",".join("?" * len(ids))
            where += f" AND dh.id IN ({placeholders})"
            params.extend(ids)
    else:
        if date_from:
            where += " AND dh.date >= ?"
            params.append(date_from)
        if date_to:
            where += " AND dh.date <= ?"
            params.append(date_to)

    return where, params, None


def _run_summary_query(db, where, params):
    """Execute the summary GROUP BY query.

    Rows are split by rate (not averaged) so the report can show — and edit —
    the exact rate carried by the underlying docket lines. line_ids lets the
    rate-review flow re-rate exactly the lines behind a row.
    """
    return db.execute(
        f"""SELECT
                COALESCE(r.category, 'Uncategorised') AS category,
                COALESCE(r.description, dl.description, 'Unknown') AS resource_desc,
                dl.resource_id,
                r.standard_rate,
                dl.unit,
                dl.rate,
                SUM(dl.qty) AS total_qty,
                SUM(dl.amount) AS subtotal,
                COUNT(DISTINCT dh.id) AS docket_count,
                GROUP_CONCAT(dl.id) AS line_ids
            FROM docket_lines dl
            JOIN docket_headers dh ON dl.docket_id = dh.id
            LEFT JOIN resources r ON dl.resource_id = r.id
            WHERE {where}
            GROUP BY category, resource_desc, dl.resource_id, dl.unit, dl.rate
            ORDER BY category, resource_desc, dl.rate""",
        params,
    ).fetchall()


@bp.route("/projects/<int:project_id>/docket-summary", methods=["GET"])
def docket_summary(project_id):
    """Docket summary grouped by category/resource for a supplier."""
    db = get_db()
    where, params, err = _build_summary_filter(project_id, request.args)
    if err:
        return jsonify({"error": err}), 400

    rows = _run_summary_query(db, where, params)

    groups = {}
    grand_total = 0
    for r in rows:
        d = dict(r)
        cat = d["category"]
        if cat not in groups:
            groups[cat] = {"category": cat, "items": [], "category_total": 0}
        groups[cat]["items"].append(d)
        groups[cat]["category_total"] += d["subtotal"] or 0
        grand_total += d["subtotal"] or 0

    return jsonify({
        "supplier": request.args.get("supplier"),
        "date_from": request.args.get("date_from"),
        "date_to": request.args.get("date_to"),
        "docket_ids": request.args.get("docket_ids"),
        "groups": list(groups.values()),
        "grand_total": grand_total,
    })


@bp.route("/projects/<int:project_id>/docket-summary/csv", methods=["GET"])
def docket_summary_csv(project_id):
    """Export docket summary as CSV."""
    db = get_db()
    where, params, err = _build_summary_filter(project_id, request.args)
    if err:
        return jsonify({"error": err}), 400

    rows = _run_summary_query(db, where, params)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Category", "Resource", "Unit", "Total Qty",
        "Rate", "Subtotal", "Docket Count",
    ])
    for r in rows:
        writer.writerow([
            r["category"], r["resource_desc"], r["unit"],
            r["total_qty"], round(r["rate"] or 0, 2),
            round(r["subtotal"], 2), r["docket_count"],
        ])

    supplier = request.args.get("supplier", "unknown")
    filename = f"docket_summary_{supplier.replace(' ', '_')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@bp.route("/projects/<int:project_id>/rerate", methods=["POST"])
def rerate_lines(project_id):
    """Re-rate docket lines from the summary report rate-review flow.

    Body:
        line_ids: [int]          — the lines behind the edited report row
        new_rate: float          — the rate read off the invoice
        resource_id: int|null    — set with update_standard to also update
                                   the resource's standard rate
        update_standard: bool
        add_resource: {description, unit, supplier_name, category} | null
                                 — create a resource from a free-text row and
                                   link these lines to it

    Returns previous values so the client can offer Undo.
    """
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    line_ids = data.get("line_ids") or []
    if not isinstance(line_ids, list) or not all(
        isinstance(i, int) for i in line_ids
    ) or not line_ids:
        return jsonify({"error": "line_ids must be a non-empty list of ids"}), 400

    try:
        new_rate = float(data.get("new_rate"))
    except (TypeError, ValueError):
        return jsonify({"error": "new_rate must be a number"}), 400
    if new_rate < 0:
        return jsonify({"error": "new_rate cannot be negative"}), 400

    # Only touch lines that belong to this project
    placeholders = ",".join("?" * len(line_ids))
    lines = db.execute(
        f"""SELECT dl.id, dl.rate FROM docket_lines dl
            JOIN docket_headers dh ON dl.docket_id = dh.id
            WHERE dl.id IN ({placeholders}) AND dh.project_id = ?""",
        (*line_ids, project_id),
    ).fetchall()
    if len(lines) != len(line_ids):
        return jsonify({"error": "One or more lines not found in this project"}), 404

    old_rate = lines[0]["rate"]

    db.execute(
        f"""UPDATE docket_lines
            SET rate = ?, amount = ROUND(qty * ?, 2)
            WHERE id IN ({placeholders})""",
        (new_rate, new_rate, *line_ids),
    )

    result = {
        "updated_lines": len(line_ids),
        "new_rate": new_rate,
        "old_rate": old_rate,
        "standard_updated": False,
        "old_standard_rate": None,
        "new_resource_id": None,
    }

    # Free-text row promoted to a resource
    add_resource = data.get("add_resource")
    if add_resource and add_resource.get("description") and add_resource.get("unit"):
        cur = db.execute(
            """INSERT INTO resources (description, unit, supplier_name, standard_rate, category)
               VALUES (?, ?, ?, ?, ?)""",
            (
                add_resource["description"],
                add_resource["unit"],
                register_supplier(db, add_resource.get("supplier_name")),
                new_rate,
                add_resource.get("category"),
            ),
        )
        new_resource_id = cur.lastrowid
        db.execute(
            f"UPDATE docket_lines SET resource_id = ? WHERE id IN ({placeholders})",
            (new_resource_id, *line_ids),
        )
        result["new_resource_id"] = new_resource_id

    # Update the resource's standard rate
    elif data.get("update_standard") and data.get("resource_id"):
        row = db.execute(
            "SELECT standard_rate FROM resources WHERE id = ?",
            (data["resource_id"],),
        ).fetchone()
        if row is not None:
            result["old_standard_rate"] = row["standard_rate"]
            db.execute(
                "UPDATE resources SET standard_rate = ?, updated_at = datetime('now') WHERE id = ?",
                (new_rate, data["resource_id"]),
            )
            result["standard_updated"] = True

    db.commit()
    return jsonify(result)


@bp.route("/projects/<int:project_id>/dockets/by-supplier", methods=["GET"])
def list_dockets_by_supplier(project_id):
    """List docket headers for a specific supplier (for docket picker)."""
    db = get_db()
    supplier = request.args.get("supplier")
    if not supplier:
        return jsonify({"error": "supplier parameter is required"}), 400

    where = "dh.project_id = ? AND dh.supplier_name = ?"
    params = [project_id, supplier]

    if request.args.get("unclaimed") == "1":
        where += " AND (dh.claimed_reference IS NULL OR dh.claimed_reference = '')"

    rows = db.execute(
        f"""SELECT dh.id, dh.date, dh.docket_number,
                  dh.claimed_reference, dh.claimed_at,
                  COALESCE(SUM(dl.amount), 0) AS total_amount,
                  COUNT(dl.id) AS line_count
           FROM docket_headers dh
           LEFT JOIN docket_lines dl ON dl.docket_id = dh.id
           WHERE {where}
           GROUP BY dh.id
           ORDER BY dh.date DESC, dh.docket_number""",
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.route("/projects/<int:project_id>/dockets/claim", methods=["POST"])
def claim_dockets(project_id):
    """Mark selected dockets as claimed with a reference string."""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
    docket_ids = data.get("docket_ids", [])
    reference = data.get("reference", "").strip()
    if not docket_ids:
        return jsonify({"error": "docket_ids is required"}), 400
    if not reference:
        return jsonify({"error": "reference is required"}), 400

    placeholders = ",".join("?" * len(docket_ids))
    db.execute(
        f"""UPDATE docket_headers
            SET claimed_reference = ?, claimed_at = datetime('now'),
                updated_at = datetime('now')
            WHERE id IN ({placeholders}) AND project_id = ?""",
        [reference] + list(docket_ids) + [project_id],
    )
    db.commit()
    return jsonify({"claimed": len(docket_ids), "reference": reference})


@bp.route("/projects/<int:project_id>/dockets/unclaim", methods=["POST"])
def unclaim_dockets(project_id):
    """Clear claim reference from selected dockets."""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
    docket_ids = data.get("docket_ids", [])
    if not docket_ids:
        return jsonify({"error": "docket_ids is required"}), 400

    placeholders = ",".join("?" * len(docket_ids))
    db.execute(
        f"""UPDATE docket_headers
            SET claimed_reference = NULL, claimed_at = NULL,
                updated_at = datetime('now')
            WHERE id IN ({placeholders}) AND project_id = ?""",
        list(docket_ids) + [project_id],
    )
    db.commit()
    return jsonify({"unclaimed": len(docket_ids)})


@bp.route("/projects/<int:project_id>/check-hashes", methods=["POST"])
def check_hashes(project_id):
    """Check which source hashes already exist. Returns matched hashes."""
    db = get_db()
    data = request.get_json()
    if not data or not isinstance(data.get("hashes"), list):
        return jsonify({"error": "hashes array is required"}), 400

    hashes = data["hashes"]
    if not hashes:
        return jsonify({"existing": []})

    placeholders = ",".join("?" * len(hashes))
    rows = db.execute(
        f"""SELECT source_hash, id, docket_number, date
            FROM docket_headers
            WHERE project_id = ? AND source_hash IN ({placeholders})""",
        [project_id] + hashes,
    ).fetchall()

    return jsonify({
        "existing": [dict(r) for r in rows],
    })


def _scan_root():
    """When DCT_OS_SCAN_ROOT is set (the hosted demo), /api/scans only serves
    files inside that directory — so a crafted docket can't turn the endpoint
    into an arbitrary-file reader. Unset on a local install: the user's own
    scans legitimately live anywhere on their machine."""
    root = os.environ.get("DCT_OS_SCAN_ROOT", "").strip()
    return os.path.abspath(root) if root else None


def _within(path, root):
    """True if `path` is inside directory `root` (both absolute)."""
    try:
        return os.path.commonpath([os.path.abspath(path), root]) == root
    except ValueError:
        return False  # e.g. different drives on Windows


@bp.route("/scans/<int:docket_id>", methods=["GET"])
def serve_scan(docket_id):
    """Serve the original scan file from the stored filepath, verifying
    the SHA-256 fingerprint still matches."""
    db = get_db()
    row = db.execute(
        "SELECT source_filepath, source_hash FROM docket_headers WHERE id = ?",
        (docket_id,),
    ).fetchone()
    if row is None:
        return jsonify({"error": "Docket not found"}), 404
    filepath = row["source_filepath"]
    expected_hash = row["source_hash"]
    if not filepath:
        return jsonify({"error": "No scan associated with this docket"}), 404
    # On the demo, refuse to serve anything outside the bundled sample-scan root.
    root = _scan_root()
    if root is not None and not _within(filepath, root):
        return jsonify({"error": "Scan not available"}), 403
    if not os.path.isfile(filepath):
        return jsonify({"error": "File not found", "path": filepath}), 404

    if expected_hash:
        sha = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                sha.update(chunk)
        if sha.hexdigest() != expected_hash:
            return jsonify({
                "error": "Fingerprint mismatch",
                "expected": expected_hash,
                "actual": sha.hexdigest(),
            }), 409

    return send_file(filepath)


@bp.route("/check-path", methods=["POST"])
def check_path():
    """Tell the folder-browse UI whether a pasted path resolves to a real file.

    Used only by the manual-entry fallback (when the native picker can't open).
    The browser sandbox never reveals the true on-disk path of a picked folder.
    """
    data = request.get_json() or {}
    path = data.get("path") or ""
    return jsonify({
        "absolute": os.path.isabs(path),
        "exists": bool(path) and os.path.isfile(path),
    })


# Folders the user has explicitly browsed this run — the only directories
# /api/scan-file will serve previews from. In-memory; cleared on restart.
_browsable_dirs = set()
_SCAN_EXTS = (".pdf", ".jpg", ".jpeg", ".png")


def _fs_browse_disabled():
    """True where the server filesystem must not be exposed to the client —
    e.g. the public hosted demo, which sets DCT_OS_NO_FS_BROWSE=1. Folder
    browsing is a local-install feature; a local user leaves this unset."""
    return os.environ.get("DCT_OS_NO_FS_BROWSE", "").strip().lower() in (
        "1", "true", "yes")


def _scan_folder(folder):
    """List + SHA-256 fingerprint the scannable files in a folder (top level,
    non-recursive). Each file comes back with its REAL absolute path so the
    docket can store a path that actually reopens later. Remembers the folder
    so /api/scan-file may preview from it."""
    files = []
    for name in sorted(os.listdir(folder)):
        if not name.lower().endswith(_SCAN_EXTS):
            continue
        full = os.path.join(folder, name)
        if not os.path.isfile(full):
            continue
        sha = hashlib.sha256()
        with open(full, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                sha.update(chunk)
        files.append({"name": name, "path": full, "hash": sha.hexdigest()})
    _browsable_dirs.add(os.path.abspath(folder))
    return files


@bp.route("/pick-folder", methods=["POST"])
def pick_folder():
    """Open a native folder picker on the local machine and return the scannable
    files inside, each with its true absolute path + fingerprint.

    DCT-OS runs on the user's own computer, so the server can show a native
    dialog and read the folder directly — the browser never needs (and the
    sandbox won't give) the real filesystem path. The dialog runs in a FRESH
    SUBPROCESS so it works reliably every time (tkinter misbehaves when reused
    across the server's worker threads — a second pick came back empty). Returns
    available=false so the UI falls back to manual entry on a headless box."""
    if _fs_browse_disabled():
        return jsonify({"available": False,
                        "error": "Folder browsing is disabled on this server"})
    import subprocess
    import sys
    import tempfile

    if getattr(sys, "frozen", False):
        cmd = [sys.executable, "pick-folder"]
    else:
        cmd = [sys.executable, "-m", "dct_os", "pick-folder"]
    fd, tmp = tempfile.mkstemp(suffix=".pick")
    os.close(fd)
    cmd.append(tmp)
    try:
        subprocess.run(cmd, timeout=600,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        with open(tmp, encoding="utf-8") as f:
            out = f.read().strip()
    except Exception as e:
        return jsonify({"available": False, "error": str(e)})
    finally:
        try:
            os.unlink(tmp)
        except Exception:
            pass

    if out == "__NO_GUI__":
        return jsonify({"available": False,
                        "error": "No folder picker on this machine — type the path"})
    if not out:
        return jsonify({"available": True, "cancelled": True})
    folder = os.path.abspath(out)
    if not os.path.isdir(folder):
        return jsonify({"available": True, "cancelled": True})
    return jsonify({"available": True, "folder": folder,
                    "files": _scan_folder(folder)})


@bp.route("/list-folder", methods=["POST"])
def list_folder():
    """Fallback for when the native picker can't open (headless/hosted): list +
    fingerprint the scannable files at a typed absolute folder path."""
    if _fs_browse_disabled():
        return jsonify({"error": "Folder browsing is disabled on this server"}), 403
    data = request.get_json() or {}
    folder = (data.get("path") or "").strip()
    if not folder:
        return jsonify({"error": "No path given"}), 400
    if not os.path.isdir(folder):
        return jsonify({"error": "Folder not found", "path": folder}), 404
    folder = os.path.abspath(folder)
    return jsonify({"folder": folder, "files": _scan_folder(folder)})


@bp.route("/scan-file", methods=["GET"])
def scan_file():
    """Serve a scan by absolute path for preview during entry — restricted to
    folders the user has browsed this run and to scan file types."""
    if _fs_browse_disabled():
        return jsonify({"error": "Folder browsing is disabled on this server"}), 403
    path = request.args.get("path", "")
    if not path:
        return jsonify({"error": "No path"}), 400
    parent = os.path.dirname(os.path.abspath(path))
    if parent not in _browsable_dirs:
        return jsonify({"error": "Path not in a browsed folder"}), 403
    if not path.lower().endswith(_SCAN_EXTS):
        return jsonify({"error": "Not a scan file"}), 400
    if not os.path.isfile(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(path)


def _docket_export_rows(db, project_id, ids_arg):
    """Rows for the docket exports (one per line item). An optional
    comma-separated `docket_ids` narrows the export to a selection — the
    grid's currently-filtered view — instead of the whole project. Blank or
    absent means export everything (the original behaviour).
    """
    sql = """SELECT dh.date, dh.docket_number, dh.supplier_name,
                  po.number AS po_number,
                  wo.number AS wo_number,
                  cc.code AS cost_code,
                  r.description AS resource_description,
                  dl.description, dl.qty, dl.unit, dl.rate, dl.amount,
                  dh.notes, dh.claimed_reference
           FROM docket_headers dh
           LEFT JOIN purchase_orders po ON dh.purchase_order_id = po.id
           JOIN docket_lines dl ON dl.docket_id = dh.id
           LEFT JOIN work_orders wo ON dl.work_order_id = wo.id
           LEFT JOIN cost_codes cc ON dl.cost_code_id = cc.id
           LEFT JOIN resources r ON dl.resource_id = r.id
           WHERE dh.project_id = ?"""
    params = [project_id]
    ids = [int(x) for x in (ids_arg or "").split(",") if x.strip().isdigit()]
    if ids:
        sql += " AND dh.id IN (%s)" % ",".join("?" * len(ids))
        params.extend(ids)
    sql += " ORDER BY dh.date, dh.docket_number, dl.sort_order"
    return db.execute(sql, params).fetchall()


@bp.route("/projects/<int:project_id>/dockets/export-csv", methods=["GET"])
def export_dockets_csv(project_id):
    """Export docket data as CSV (one row per line item). Honours an optional
    docket_ids selection."""
    db = get_db()
    rows = _docket_export_rows(db, project_id, request.args.get("docket_ids"))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date", "Docket #", "Supplier", "PO #",
        "WO #", "Cost Code", "Resource", "Description",
        "Qty", "Unit", "Rate", "Amount", "Notes", "Claimed",
    ])
    for r in rows:
        writer.writerow([
            r["date"], r["docket_number"], r["supplier_name"],
            r["po_number"] or "", r["wo_number"] or "",
            r["cost_code"] or "", r["resource_description"] or "",
            r["description"] or "", r["qty"], r["unit"] or "",
            round(r["rate"] or 0, 2), round(r["amount"] or 0, 2),
            r["notes"] or "", r["claimed_reference"] or "",
        ])

    project = db.execute(
        "SELECT code FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    code = (project["code"] or "project").replace(" ", "_") if project else "project"
    filename = f"dockets_{code}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_workbook(title, headers, currency_cols):
    """Create a styled openpyxl workbook with a header row. Returns (wb, ws)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws._dct_currency_cols = currency_cols
    return wb, ws


def _xlsx_finish(wb, ws, col_widths, filename):
    """Apply widths + currency formats and return the download response."""
    from openpyxl.utils import get_column_letter

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for col in getattr(ws, "_dct_currency_cols", []):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@bp.route("/projects/<int:project_id>/dockets/export-xlsx", methods=["GET"])
def export_dockets_xlsx(project_id):
    """Export dockets as a formatted Excel workbook (one row per line).
    Honours an optional docket_ids selection."""
    db = get_db()
    rows = _docket_export_rows(db, project_id, request.args.get("docket_ids"))

    wb, ws = _xlsx_workbook(
        "Dockets",
        ["Date", "Docket #", "Supplier", "PO #", "WO #", "Cost Code",
         "Resource", "Description", "Qty", "Unit", "Rate", "Amount",
         "Notes", "Claimed"],
        currency_cols=[11, 12],
    )
    for r in rows:
        ws.append([
            r["date"], r["docket_number"], r["supplier_name"],
            r["po_number"] or "", r["wo_number"] or "",
            r["cost_code"] or "", r["resource_description"] or "",
            r["description"] or "", r["qty"], r["unit"] or "",
            round(r["rate"] or 0, 2), round(r["amount"] or 0, 2),
            r["notes"] or "", r["claimed_reference"] or "",
        ])

    project = db.execute(
        "SELECT code FROM projects WHERE id = ?", (project_id,)
    ).fetchone()
    code = (project["code"] or "project").replace(" ", "_") if project else "project"
    return _xlsx_finish(
        wb, ws,
        [11, 13, 22, 11, 11, 12, 26, 30, 8, 7, 10, 12, 24, 12],
        f"dockets_{code}.xlsx",
    )


@bp.route("/projects/<int:project_id>/docket-summary/xlsx", methods=["GET"])
def docket_summary_xlsx(project_id):
    """Export the docket summary as a formatted Excel workbook."""
    from openpyxl.styles import Font

    db = get_db()
    where, params, err = _build_summary_filter(project_id, request.args)
    if err:
        return jsonify({"error": err}), 400

    rows = _run_summary_query(db, where, params)

    wb, ws = _xlsx_workbook(
        "Docket Summary",
        ["Category", "Resource", "Unit", "Total Qty", "Rate", "Subtotal",
         "Docket Count"],
        currency_cols=[5, 6],
    )
    grand_total = 0
    for r in rows:
        ws.append([
            r["category"], r["resource_desc"], r["unit"],
            r["total_qty"], round(r["rate"] or 0, 2),
            round(r["subtotal"], 2), r["docket_count"],
        ])
        grand_total += r["subtotal"] or 0

    ws.append(["", "", "", "", "Grand Total", round(grand_total, 2), ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    supplier = request.args.get("supplier", "unknown").replace(" ", "_")
    return _xlsx_finish(
        wb, ws,
        [18, 30, 8, 10, 10, 12, 12],
        f"docket_summary_{supplier}.xlsx",
    )


@bp.route("/projects/<int:project_id>/dockets/import-csv", methods=["POST"])
def import_dockets_csv(project_id):
    """Import dockets from CSV. Matches WO/CC/Resource by code/number."""
    db = get_db()

    if "file" not in request.files:
        # Try JSON body with csv_text
        data = request.get_json()
        if not data or not data.get("csv_text"):
            return jsonify({"error": "CSV file or csv_text required"}), 400
        reader = csv.DictReader(io.StringIO(data["csv_text"]))
    else:
        f = request.files["file"]
        text = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

    # Build lookup maps for this project
    wos = {
        r["number"]: r["id"]
        for r in db.execute(
            "SELECT id, number FROM work_orders WHERE project_id = ?",
            (project_id,),
        ).fetchall()
    }
    ccs = {
        r["code"]: r["id"]
        for r in db.execute(
            "SELECT id, code FROM cost_codes WHERE project_id = ?",
            (project_id,),
        ).fetchall()
    }
    pos = {
        r["number"]: r["id"]
        for r in db.execute(
            "SELECT id, number FROM purchase_orders WHERE project_id = ?",
            (project_id,),
        ).fetchall()
    }
    resources = {
        r["description"].lower(): r["id"]
        for r in db.execute("SELECT id, description FROM resources").fetchall()
    }

    # Group CSV rows into docket headers
    groups = {}
    row_count = 0
    for row in reader:
        row_count += 1
        date = (row.get("Date") or "").strip()
        docket_num = (row.get("Docket #") or "").strip()
        supplier = (row.get("Supplier") or "").strip()
        if not date:
            continue

        key = (date, docket_num, supplier)
        if key not in groups:
            groups[key] = {
                "date": date,
                "docket_number": docket_num or None,
                "supplier_name": supplier or None,
                "po_number": (row.get("PO #") or "").strip() or None,
                "notes": (row.get("Notes") or "").strip() or None,
                "lines": [],
            }

        wo_num = (row.get("WO #") or "").strip()
        cc_code = (row.get("Cost Code") or "").strip()
        res_desc = (row.get("Resource") or "").strip()

        try:
            qty = float(row.get("Qty") or 0)
        except (ValueError, TypeError):
            qty = 0
        try:
            rate = float(row.get("Rate") or 0)
        except (ValueError, TypeError):
            rate = 0

        groups[key]["lines"].append({
            "work_order_id": wos.get(wo_num),
            "cost_code_id": ccs.get(cc_code),
            "resource_id": resources.get(res_desc.lower()) if res_desc else None,
            "description": (row.get("Description") or "").strip() or None,
            "qty": qty,
            "unit": (row.get("Unit") or "").strip() or None,
            "rate": rate,
        })

    # Insert dockets, skipping duplicates
    created = 0
    skipped = 0
    for key, grp in groups.items():
        date, docket_num, supplier = key
        if docket_num and supplier:
            dup = db.execute(
                """SELECT id FROM docket_headers
                   WHERE project_id = ? AND supplier_name = ?
                   AND docket_number = ? AND date = ?""",
                (project_id, supplier, docket_num, date),
            ).fetchone()
            if dup:
                skipped += 1
                continue

        po_id = pos.get(grp["po_number"]) if grp["po_number"] else None

        cur = db.execute(
            """INSERT INTO docket_headers
               (project_id, purchase_order_id, supplier_name, date,
                docket_number, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (project_id, po_id, register_supplier(db, grp["supplier_name"]), grp["date"],
             grp["docket_number"], grp["notes"]),
        )
        header_id = cur.lastrowid

        for i, line in enumerate(grp["lines"]):
            amount = line["qty"] * line["rate"]
            db.execute(
                """INSERT INTO docket_lines
                   (docket_id, work_order_id, cost_code_id, resource_id,
                    description, qty, unit, rate, amount, sort_order)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (header_id, line["work_order_id"], line["cost_code_id"],
                 line["resource_id"], line["description"],
                 line["qty"], line["unit"], line["rate"], amount, i),
            )
        created += 1

    db.commit()
    return jsonify({
        "created": created,
        "skipped": skipped,
        "rows_read": row_count,
    })


@bp.route("/projects/<int:project_id>/check-duplicate", methods=["GET"])
def check_duplicate(project_id):
    """Check if a docket with same supplier+number+date already exists."""
    db = get_db()
    supplier = request.args.get("supplier")
    docket_number = request.args.get("docket_number")
    date = request.args.get("date")
    exclude_id = request.args.get("exclude_id")

    if not supplier or not docket_number or not date:
        return jsonify({"duplicate": False})

    params = [project_id, supplier, docket_number, date]
    where = """project_id = ? AND supplier_name = ?
               AND docket_number = ? AND date = ?"""
    if exclude_id:
        where += " AND id != ?"
        params.append(int(exclude_id))

    row = db.execute(
        f"SELECT id FROM docket_headers WHERE {where} LIMIT 1",
        params,
    ).fetchone()

    return jsonify({
        "duplicate": row is not None,
        "existing_id": row["id"] if row else None,
    })
