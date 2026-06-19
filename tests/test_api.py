import json
import os
import tempfile

import pytest

from dct_os.app import create_app


@pytest.fixture
def client():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        app = create_app({"TESTING": True, "DATABASE": path})
        with app.test_client() as client:
            yield client
    finally:
        os.unlink(path)


def _create_test_dockets(client):
    """Create a self-contained test project with WOs, CCs, PO, and dockets.

    Creates:
      - 1 project ("Test Project", TEST-001)
      - 2 work orders (WT-001 Earthworks, WT-002 Pavement)
      - 2 cost codes (CT-01 Earthworks $420k, CT-02 Pavement $650k)
      - 1 purchase order (PT-001 Test Supplier $45,000)
      - 3 docket headers:
          RGC-T001 (2025-02-03): 2 lines, total $2,640
          RGC-T003 (2025-02-04): 1 line,  total $1,125
          RGC-T005 (2025-02-05): 1 line,  total $1,880
      Grand total: $5,645

    Returns (project_id, [docket_id1, docket_id2, docket_id3]).
    """
    # Create test project
    resp = client.post("/api/projects", json={
        "name": "Test Project", "code": "TEST-001", "status": "Active",
    }, content_type="application/json")
    project = json.loads(resp.data)
    pid = project["id"]

    # Create work orders
    wo1 = json.loads(client.post(
        f"/api/projects/{pid}/work-orders",
        json={"number": "WT-001", "description": "Earthworks", "status": "Active"},
        content_type="application/json",
    ).data)
    wo2 = json.loads(client.post(
        f"/api/projects/{pid}/work-orders",
        json={"number": "WT-002", "description": "Pavement", "status": "Active"},
        content_type="application/json",
    ).data)

    # Create cost codes
    cc1 = json.loads(client.post(
        f"/api/projects/{pid}/cost-codes",
        json={"code": "CT-01", "description": "Earthworks", "budget_amount": 420000},
        content_type="application/json",
    ).data)
    cc2 = json.loads(client.post(
        f"/api/projects/{pid}/cost-codes",
        json={"code": "CT-02", "description": "Pavement", "budget_amount": 650000},
        content_type="application/json",
    ).data)

    # Create purchase order
    po = json.loads(client.post(
        f"/api/projects/{pid}/purchase-orders",
        json={"number": "PT-001", "supplier_name": "Test Supplier",
              "value": 45000, "is_active": 1},
        content_type="application/json",
    ).data)

    # Create 3 dockets
    dockets = [
        {
            "date": "2025-02-03",
            "supplier_name": "Test Supplier",
            "docket_number": "RGC-T001",
            "purchase_order_id": po["id"],
            "lines": [
                {"work_order_id": wo1["id"], "cost_code_id": cc1["id"],
                 "description": "Earthworks - 20T Exc", "qty": 8, "unit": "Hr", "rate": 220},
                {"work_order_id": wo1["id"], "cost_code_id": cc1["id"],
                 "description": "Earthworks - Super", "qty": 8, "unit": "Hr", "rate": 110},
            ],
        },
        {
            "date": "2025-02-04",
            "supplier_name": "Test Supplier",
            "docket_number": "RGC-T003",
            "purchase_order_id": po["id"],
            "lines": [
                {"work_order_id": wo1["id"], "cost_code_id": cc1["id"],
                 "description": "Earthworks - 10T Tip", "qty": 9, "unit": "Hr", "rate": 125},
            ],
        },
        {
            "date": "2025-02-05",
            "supplier_name": "Test Supplier",
            "docket_number": "RGC-T005",
            "purchase_order_id": po["id"],
            "lines": [
                {"work_order_id": wo2["id"], "cost_code_id": cc2["id"],
                 "description": "Pavement - Grader", "qty": 8, "unit": "Hr", "rate": 235},
            ],
        },
    ]
    ids = []
    for d in dockets:
        resp = client.post(f"/api/projects/{pid}/dockets", json=d,
                           content_type="application/json")
        ids.append(json.loads(resp.data)["id"])
    return pid, ids


# ---------------------------------------------------------------------------
# Static data tests (projects, cost codes, resources, work orders, POs)
# These rely only on seed.sql reference data -- no dockets needed.
# ---------------------------------------------------------------------------

def test_index(client):
    resp = client.get("/")
    assert resp.status_code == 200


def test_list_projects(client):
    resp = client.get("/api/projects")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) >= 1  # seed creates 1 project
    names = [p["name"] for p in data]
    assert "Warrawong Road Rehabilitation" in names


def test_create_project(client):
    resp = client.post(
        "/api/projects",
        data=json.dumps({"name": "New Test Project", "code": "NTP-001"}),
        content_type="application/json",
    )
    assert resp.status_code == 201
    data = json.loads(resp.data)
    assert data["name"] == "New Test Project"
    assert data["code"] == "NTP-001"
    assert "id" in data


def test_update_project(client):
    resp = client.put(
        "/api/projects/1",
        data=json.dumps({"name": "Updated Project Name"}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["name"] == "Updated Project Name"


def test_delete_project(client):
    # Create a project to delete (seed only has 1 project now)
    create = client.post(
        "/api/projects",
        data=json.dumps({"name": "To Delete", "code": "DEL-001"}),
        content_type="application/json",
    )
    pid = json.loads(create.data)["id"]
    resp = client.delete(f"/api/projects/{pid}")
    assert resp.status_code == 200
    check = client.get(f"/api/projects/{pid}")
    assert check.status_code == 404


def test_cost_codes(client):
    resp = client.get("/api/projects/1/cost-codes")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 10
    first = data[0]
    assert "code" in first
    assert "description" in first
    assert "budget_amount" in first


def test_resources(client):
    resp = client.get("/api/resources")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 48
    first = data[0]
    assert "description" in first
    assert "unit" in first
    assert "details" in first
    assert "standard_rate" in first


def test_work_orders(client):
    resp = client.get("/api/projects/1/work-orders")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 7
    first = data[0]
    assert "number" in first
    assert "description" in first
    assert "status" in first


def test_create_work_order(client):
    resp = client.post(
        "/api/projects/1/work-orders",
        data=json.dumps({"number": "W-NEW-001", "description": "New WO"}),
        content_type="application/json",
    )
    assert resp.status_code == 201
    data = json.loads(resp.data)
    assert data["number"] == "W-NEW-001"


def test_create_purchase_order(client):
    resp = client.post(
        "/api/projects/1/purchase-orders",
        data=json.dumps({
            "number": "99999",
            "supplier_name": "Test Supplier",
            "value": 10000.00,
            "is_active": 1,
        }),
        content_type="application/json",
    )
    assert resp.status_code == 201
    data = json.loads(resp.data)
    assert data["number"] == "99999"
    assert data["value"] == 10000.00


def test_wo_cost_codes_matrix(client):
    resp = client.get("/api/work-orders/1/cost-codes")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) >= 1


def test_po_work_orders(client):
    resp = client.get("/api/purchase-orders/1/work-orders")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) >= 1


def test_404_on_missing(client):
    resp = client.get("/api/projects/999")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Docket CRUD tests -- create test dockets via API, then exercise endpoints.
# All docket tests use a self-contained test project to avoid seed conflicts.
# ---------------------------------------------------------------------------

def test_dockets_for_project(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/dockets")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 3
    first = data[0]
    assert "lines" in first
    assert "total_amount" in first
    assert "line_count" in first
    assert "po_number" in first
    assert "wo_numbers" in first
    assert first["line_count"] == len(first["lines"])
    assert first["total_amount"] > 0


def test_dockets_empty_project(client):
    """A project with no dockets returns an empty list, not an error."""
    # Project 2 (Myall Creek) has no seed dockets
    resp = client.get("/api/projects/2/dockets")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data == []


def test_get_single_docket(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/dockets/{ids[0]}")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["id"] == ids[0]
    assert "lines" in data
    assert len(data["lines"]) == 2
    assert data["lines"][0]["wo_number"] is not None
    assert data["lines"][0]["cost_code"] is not None


def test_create_docket_with_lines(client):
    resp = client.post(
        "/api/projects/1/dockets",
        data=json.dumps({
            "date": "2025-05-01",
            "supplier_name": "Test Supplier",
            "docket_number": "TST-001",
            "purchase_order_id": 1,
            "lines": [
                {
                    "work_order_id": 1,
                    "cost_code_id": 1,
                    "description": "Line A",
                    "qty": 8.0,
                    "unit": "Hr",
                    "rate": 100.00,
                },
                {
                    "work_order_id": 2,
                    "cost_code_id": 2,
                    "description": "Line B",
                    "qty": 4.0,
                    "unit": "Hr",
                    "rate": 200.00,
                },
            ],
        }),
        content_type="application/json",
    )
    assert resp.status_code == 201
    data = json.loads(resp.data)
    assert data["supplier_name"] == "Test Supplier"
    assert data["purchase_order_id"] == 1
    assert data["line_count"] == 2
    assert data["total_amount"] == 1600.0
    assert data["lines"][0]["amount"] == 800.0
    assert data["lines"][1]["amount"] == 800.0


def test_update_docket_header(client):
    pid, ids = _create_test_dockets(client)
    resp = client.put(
        f"/api/dockets/{ids[0]}",
        data=json.dumps({"supplier_name": "Updated Supplier"}),
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["supplier_name"] == "Updated Supplier"
    assert len(data["lines"]) == 2


def test_update_docket_lines(client):
    pid, ids = _create_test_dockets(client)
    resp = client.put(
        f"/api/dockets/{ids[0]}",
        data=json.dumps({
            "lines": [
                {
                    "work_order_id": 1,
                    "cost_code_id": 1,
                    "description": "Replaced line",
                    "qty": 10.0,
                    "unit": "Hr",
                    "rate": 50.00,
                },
            ],
        }),
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["line_count"] == 1
    assert data["total_amount"] == 500.0
    assert data["lines"][0]["description"] == "Replaced line"


def test_delete_docket(client):
    pid, ids = _create_test_dockets(client)
    resp = client.delete(f"/api/dockets/{ids[0]}")
    assert resp.status_code == 200
    check = client.get(f"/api/dockets/{ids[0]}")
    assert check.status_code == 404


# ---------------------------------------------------------------------------
# Summary / reporting tests -- need docket data to produce meaningful output.
# ---------------------------------------------------------------------------

def test_project_summary(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/summary")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["total_dockets"] == 3
    assert data["total_spend"] == 5645.0  # 2640 + 1125 + 1880
    assert data["supplier_count"] == 1


def test_project_summary_empty(client):
    """Summary with no dockets returns zeroes, not an error."""
    # Project 2 has no seed dockets
    resp = client.get("/api/projects/2/summary")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["total_dockets"] == 0
    assert data["total_spend"] == 0


def test_purchase_orders_with_drawdown(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/purchase-orders")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 1  # test project has 1 PO
    po = data[0]
    assert po["number"] == "PT-001"
    assert po["value"] == 45000.00
    assert po["spent"] == 5645.0
    assert po["remaining"] == 45000.00 - 5645.0


def test_cost_report(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/cost-report")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 2  # test project has 2 cost codes
    earthworks = next(c for c in data if c["code"] == "CT-01")
    assert earthworks["actual_spend"] == 3765.0  # 2640 + 1125
    assert earthworks["budget_amount"] == 420000
    pavement = next(c for c in data if c["code"] == "CT-02")
    assert pavement["actual_spend"] == 1880.0


def test_cost_report_empty(client):
    """Cost report with no dockets shows zero actuals."""
    # Create a fresh project with cost codes but no dockets
    proj = client.post(
        "/api/projects",
        data=json.dumps({"name": "Empty Report Test", "code": "ERT-001"}),
        content_type="application/json",
    )
    pid = json.loads(proj.data)["id"]
    # Add 2 cost codes
    for code in ("CC-A", "CC-B"):
        client.post(
            f"/api/projects/{pid}/cost-codes",
            data=json.dumps({"code": code, "description": "Test", "budget_amount": 1000}),
            content_type="application/json",
        )
    resp = client.get(f"/api/projects/{pid}/cost-report")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 2
    for cc in data:
        assert cc["actual_spend"] == 0


def test_dashboard(client):
    """The combined dashboard endpoint rolls up this project's own data."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/dashboard")
    assert resp.status_code == 200
    data = json.loads(resp.data)

    # All seven sections are present.
    for key in ("tiles", "cost_codes", "spend_series", "wo_costs",
                "po_drawdown", "top_suppliers", "claimed", "to_claim"):
        assert key in data

    # Tiles: budget = 420k + 650k, spent = 5645, 3 dockets, 1 supplier.
    tiles = data["tiles"]
    assert tiles["total_budget"] == 1070000
    assert tiles["total_spent"] == 5645.0
    assert tiles["remaining"] == 1070000 - 5645.0
    assert tiles["total_dockets"] == 3
    assert tiles["supplier_count"] == 1

    # Cost codes: same actuals as the cost report.
    ccs = {c["code"]: c for c in data["cost_codes"]}
    assert ccs["CT-01"]["actual_spend"] == 3765.0
    assert ccs["CT-02"]["actual_spend"] == 1880.0

    # Cost by work order: WT-001 = 2640 + 1125, WT-002 = 1880.
    wos = {w["number"]: w["amount"] for w in data["wo_costs"]}
    assert wos["WT-001"] == 3765.0
    assert wos["WT-002"] == 1880.0
    # Sorted descending by amount.
    amounts = [w["amount"] for w in data["wo_costs"]]
    assert amounts == sorted(amounts, reverse=True)

    # Spend over time: cumulative, ending at the grand total.
    assert data["spend_series"]
    assert data["spend_series"][-1]["cumulative"] == 5645.0

    # PO drawdown: the one active PO, committed 45k, drawn 5645.
    assert len(data["po_drawdown"]) == 1
    po = data["po_drawdown"][0]
    assert po["number"] == "PT-001"
    assert po["committed"] == 45000.0
    assert po["drawn"] == 5645.0

    # Top suppliers: one supplier with the full spend.
    assert data["top_suppliers"] == [{"name": "Test Supplier", "amount": 5645.0}]

    # Nothing claimed yet → all to-claim.
    assert data["claimed"] == 0
    assert data["to_claim"] == 5645.0


def test_dashboard_claimed_split(client):
    """Claiming a docket moves its spend from to-claim to claimed."""
    pid, ids = _create_test_dockets(client)
    # Claim the first docket (RGC-T001 = $2,640).
    resp = client.post(
        f"/api/projects/{pid}/dockets/claim",
        json={"docket_ids": [ids[0]], "reference": "CLAIM-01"},
        content_type="application/json",
    )
    assert resp.status_code == 200

    data = json.loads(client.get(f"/api/projects/{pid}/dashboard").data)
    assert data["claimed"] == 2640.0
    assert data["to_claim"] == 3005.0  # 5645 - 2640


def test_dashboard_empty_project(client):
    """A project with no dockets returns zeroed tiles and empty sections."""
    resp = client.get("/api/projects/2/dashboard")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["tiles"]["total_spent"] == 0
    assert data["tiles"]["total_dockets"] == 0
    assert data["spend_series"] == []
    assert data["wo_costs"] == []
    assert data["top_suppliers"] == []
    assert data["claimed"] == 0
    assert data["to_claim"] == 0


def test_project_suppliers(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/suppliers")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert isinstance(data, list)
    assert len(data) > 0
    # Should be sorted case-insensitively
    assert data == sorted(data, key=str.lower)
    # Test Supplier should be in the list (from dockets + POs)
    assert "Test Supplier" in data


# ---------------------------------------------------------------------------
# Docket summary report tests
# ---------------------------------------------------------------------------

def test_docket_summary(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/docket-summary?supplier=Test+Supplier")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["supplier"] == "Test Supplier"
    assert "groups" in data
    assert len(data["groups"]) > 0
    assert data["grand_total"] > 0
    # Each group has items with expected fields
    first_group = data["groups"][0]
    assert "category" in first_group
    assert "items" in first_group
    assert "category_total" in first_group
    first_item = first_group["items"][0]
    assert "resource_desc" in first_item
    assert "total_qty" in first_item
    assert "subtotal" in first_item


def test_docket_summary_date_filter(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/docket-summary?"
        "supplier=Test+Supplier&date_from=2025-02-01&date_to=2025-02-28"
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["date_from"] == "2025-02-01"
    assert data["date_to"] == "2025-02-28"
    # All 3 test dockets fall within this range
    assert data["grand_total"] == 5645.0


def test_docket_summary_requires_supplier(client):
    resp = client.get("/api/projects/1/docket-summary")
    assert resp.status_code == 400
    data = json.loads(resp.data)
    assert "supplier" in data["error"].lower()


def test_docket_summary_csv(client):
    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/docket-summary/csv?supplier=Test+Supplier"
    )
    assert resp.status_code == 200
    assert "text/csv" in resp.content_type
    text = resp.data.decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) >= 2  # header + at least one data row
    assert "Category" in lines[0]
    assert "Subtotal" in lines[0]


def test_docket_summary_by_ids(client):
    """Summary filtered by specific docket IDs instead of date range."""
    pid, ids = _create_test_dockets(client)
    # Only include first two dockets
    id_str = f"{ids[0]},{ids[1]}"
    resp = client.get(
        f"/api/projects/{pid}/docket-summary?"
        f"supplier=Test+Supplier&docket_ids={id_str}"
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["docket_ids"] == id_str
    assert data["grand_total"] == 3765.0  # 2640 + 1125 (first two dockets)

    # Unfiltered should include all three
    all_resp = client.get(
        f"/api/projects/{pid}/docket-summary?supplier=Test+Supplier"
    )
    all_data = json.loads(all_resp.data)
    assert all_data["grand_total"] == 5645.0
    assert data["grand_total"] < all_data["grand_total"]


# ---------------------------------------------------------------------------
# By-supplier and duplicate detection
# ---------------------------------------------------------------------------

def test_dockets_by_supplier(client):
    """List docket headers for a specific supplier."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/dockets/by-supplier?supplier=Test+Supplier"
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert isinstance(data, list)
    assert len(data) == 3
    first = data[0]
    assert "id" in first
    assert "date" in first
    assert "docket_number" in first
    assert "total_amount" in first


def test_dockets_by_supplier_requires_supplier(client):
    resp = client.get("/api/projects/1/dockets/by-supplier")
    assert resp.status_code == 400


def test_check_hashes(client):
    """Create a docket with source_hash, then verify check-hashes finds it."""
    pid, ids = _create_test_dockets(client)
    # Create a docket with a source hash
    docket = {
        "date": "2025-06-01",
        "supplier_name": "Test Supplier",
        "docket_number": "HASH-001",
        "source_hash": "abc123def456",
        "source_filename": "test_docket.pdf",
        "source_filepath": "/scans/test_docket.pdf",
        "lines": [{"description": "Test", "qty": 1, "rate": 100}],
    }
    resp = client.post(f"/api/projects/{pid}/dockets",
                       json=docket, content_type="application/json")
    assert resp.status_code == 201
    created = json.loads(resp.data)
    assert created["source_hash"] == "abc123def456"
    assert created["source_filename"] == "test_docket.pdf"
    assert created["source_filepath"] == "/scans/test_docket.pdf"

    # Check that the hash is found
    check = client.post(f"/api/projects/{pid}/check-hashes",
                        json={"hashes": ["abc123def456", "unknown_hash"]},
                        content_type="application/json")
    assert check.status_code == 200
    check_data = json.loads(check.data)
    existing_hashes = [e["source_hash"] for e in check_data["existing"]]
    assert "abc123def456" in existing_hashes
    assert "unknown_hash" not in existing_hashes


def test_check_duplicate(client):
    """Check duplicate detection by supplier+number+date."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/check-duplicate?"
        "supplier=Test+Supplier&docket_number=RGC-T001&date=2025-02-03"
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["duplicate"] is True
    assert data["existing_id"] is not None

    # Non-existing combo
    resp2 = client.get(
        f"/api/projects/{pid}/check-duplicate?"
        "supplier=Nobody&docket_number=NOPE-999&date=2099-01-01"
    )
    data2 = json.loads(resp2.data)
    assert data2["duplicate"] is False


# ---------------------------------------------------------------------------
# Claim / unclaim tests
# ---------------------------------------------------------------------------

def test_claim_dockets(client):
    """Claim dockets with a reference string."""
    pid, ids = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets/claim",
        json={"docket_ids": [ids[0], ids[1]], "reference": "INV-TEST-001"},
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["claimed"] == 2
    assert data["reference"] == "INV-TEST-001"

    # Verify the docket is now claimed
    docket = client.get(f"/api/dockets/{ids[0]}")
    d = json.loads(docket.data)
    assert d["claimed_reference"] == "INV-TEST-001"
    assert d["claimed_at"] is not None


def test_unclaim_dockets(client):
    """Unclaim previously claimed dockets."""
    pid, ids = _create_test_dockets(client)
    # First claim
    client.post(
        f"/api/projects/{pid}/dockets/claim",
        json={"docket_ids": [ids[0]], "reference": "INV-X"},
    )
    # Then unclaim
    resp = client.post(
        f"/api/projects/{pid}/dockets/unclaim",
        json={"docket_ids": [ids[0]]},
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["unclaimed"] == 1

    # Verify cleared
    docket = client.get(f"/api/dockets/{ids[0]}")
    d = json.loads(docket.data)
    assert d["claimed_reference"] is None
    assert d["claimed_at"] is None


def test_claim_requires_reference(client):
    pid, ids = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets/claim",
        json={"docket_ids": [ids[0]], "reference": ""},
    )
    assert resp.status_code == 400


def test_claim_requires_docket_ids(client):
    resp = client.post(
        "/api/projects/1/dockets/claim",
        json={"docket_ids": [], "reference": "INV-001"},
    )
    assert resp.status_code == 400


def test_dockets_by_supplier_unclaimed_filter(client):
    """Unclaimed filter hides claimed dockets."""
    pid, ids = _create_test_dockets(client)

    # Baseline: 3 Test Supplier dockets
    all_resp = client.get(
        f"/api/projects/{pid}/dockets/by-supplier?supplier=Test+Supplier"
    )
    all_data = json.loads(all_resp.data)
    assert len(all_data) == 3

    # Claim first docket
    client.post(
        f"/api/projects/{pid}/dockets/claim",
        json={"docket_ids": [ids[0]], "reference": "TEST-FILTER"},
    )

    # Unclaimed filter should return 2
    filtered_resp = client.get(
        f"/api/projects/{pid}/dockets/by-supplier?"
        "supplier=Test+Supplier&unclaimed=1"
    )
    filtered_data = json.loads(filtered_resp.data)
    assert len(filtered_data) == 2

    # Without filter, all 3 still returned
    unfiltered_resp = client.get(
        f"/api/projects/{pid}/dockets/by-supplier?supplier=Test+Supplier"
    )
    unfiltered_data = json.loads(unfiltered_resp.data)
    assert len(unfiltered_data) == 3


def test_dockets_by_supplier_returns_claim_fields(client):
    """By-supplier endpoint returns claimed_reference and claimed_at."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/dockets/by-supplier?supplier=Test+Supplier"
    )
    data = json.loads(resp.data)
    first = data[0]
    assert "claimed_reference" in first
    assert "claimed_at" in first


# ---------------------------------------------------------------------------
# CSV export / import tests
# ---------------------------------------------------------------------------

def test_export_dockets_csv(client):
    """Export dockets as CSV with one row per line item."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/dockets/export-csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.content_type
    text = resp.data.decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) == 5  # header + 4 line items (2+1+1)
    assert "Date" in lines[0]
    assert "Docket #" in lines[0]
    assert "Amount" in lines[0]


def test_export_dockets_csv_filtered_by_ids(client):
    """Export honours a docket_ids selection (the grid's filtered view)."""
    pid, ids = _create_test_dockets(client)
    # Export only the first docket (RGC-T001, 2 lines)
    resp = client.get(f"/api/projects/{pid}/dockets/export-csv?docket_ids={ids[0]}")
    assert resp.status_code == 200
    lines = resp.data.decode("utf-8").strip().split("\n")
    assert len(lines) == 3  # header + 2 line items, not all 4

    # Two dockets selected → their combined lines (2 + 1 = 3)
    resp = client.get(
        f"/api/projects/{pid}/dockets/export-csv?docket_ids={ids[0]},{ids[1]}"
    )
    lines = resp.data.decode("utf-8").strip().split("\n")
    assert len(lines) == 4  # header + 3 line items

    # No docket_ids → everything (4 line items across 3 dockets)
    resp = client.get(f"/api/projects/{pid}/dockets/export-csv")
    lines = resp.data.decode("utf-8").strip().split("\n")
    assert len(lines) == 5


def test_export_dockets_xlsx_filtered_by_ids(client):
    import io as _io
    from openpyxl import load_workbook

    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/dockets/export-xlsx?docket_ids={ids[0]}")
    assert resp.status_code == 200
    ws = load_workbook(_io.BytesIO(resp.data)).active
    assert ws.max_row == 3  # header + 2 lines


def test_export_empty_project_csv(client):
    """Exporting CSV from a project with no dockets returns header only."""
    resp = client.get("/api/projects/2/dockets/export-csv")
    assert resp.status_code == 200
    text = resp.data.decode("utf-8")
    lines = text.strip().split("\n")
    assert len(lines) == 1  # header only


def test_import_dockets_csv(client):
    """Import dockets from CSV text, matching WO/CC by code."""
    pid, ids = _create_test_dockets(client)

    csv_text = (
        "Date,Docket #,Supplier,PO #,WO #,Cost Code,Resource,Description,"
        "Qty,Unit,Rate,Amount,Notes,Claimed\n"
        "2025-03-01,IMP-001,Import Supplier,PT-001,WT-001,CT-01,,"
        "Imported item,5,Hr,100,500,,\n"
        "2025-03-01,IMP-001,Import Supplier,PT-001,WT-002,CT-02,,"
        "Imported item 2,3,Hr,200,600,,\n"
    )
    resp = client.post(
        f"/api/projects/{pid}/dockets/import-csv",
        json={"csv_text": csv_text},
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["created"] == 1  # 2 CSV rows = 1 docket (same date+number+supplier)
    assert data["skipped"] == 0
    assert data["rows_read"] == 2

    # Verify docket was created
    dockets = json.loads(client.get(f"/api/projects/{pid}/dockets").data)
    assert len(dockets) == 4  # 3 original + 1 imported
    imported = [d for d in dockets if d["docket_number"] == "IMP-001"]
    assert len(imported) == 1
    assert imported[0]["line_count"] == 2


def test_import_csv_skips_duplicates(client):
    """Import skips dockets with matching supplier+number+date."""
    pid, ids = _create_test_dockets(client)

    # Try importing a docket that already exists
    csv_text = (
        "Date,Docket #,Supplier,PO #,WO #,Cost Code,Resource,Description,"
        "Qty,Unit,Rate,Amount,Notes,Claimed\n"
        "2025-02-03,RGC-T001,Test Supplier,,,,,"
        "Duplicate item,1,Hr,100,100,,\n"
    )
    resp = client.post(
        f"/api/projects/{pid}/dockets/import-csv",
        json={"csv_text": csv_text},
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["created"] == 0
    assert data["skipped"] == 1

    # Docket count unchanged
    dockets = json.loads(client.get(f"/api/projects/{pid}/dockets").data)
    assert len(dockets) == 3


def test_import_csv_roundtrip(client):
    """Export then re-import should produce no new dockets (all duplicates)."""
    pid, ids = _create_test_dockets(client)

    # Export
    export_resp = client.get(f"/api/projects/{pid}/dockets/export-csv")
    csv_text = export_resp.data.decode("utf-8")

    # Re-import the same CSV
    resp = client.post(
        f"/api/projects/{pid}/dockets/import-csv",
        json={"csv_text": csv_text},
        content_type="application/json",
    )
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["created"] == 0
    assert data["skipped"] == 3  # all 3 dockets already exist


# ---------------------------------------------------------------------------
# Seed data sanity checks
# ---------------------------------------------------------------------------

def test_seed_dockets_exist_for_project_1(client):
    """Project 1 (Warrawong Road) has seed demo dockets."""
    resp = client.get("/api/projects/1/dockets")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 18  # 18 seed docket headers
    # Check a known docket
    rgc_0001 = [d for d in data if d["docket_number"] == "RGC-0001"]
    assert len(rgc_0001) == 1
    assert rgc_0001[0]["supplier_name"] == "Redgum Civil Pty Ltd"


def test_seed_dockets_absent_for_project_2(client):
    """Project 2 (Myall Creek) has no seed dockets."""
    resp = client.get("/api/projects/2/dockets")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert len(data) == 0


def test_seed_summary_project_1(client):
    """Seed docket summary for project 1 has correct totals."""
    resp = client.get("/api/projects/1/summary")
    assert resp.status_code == 200
    data = json.loads(resp.data)
    assert data["total_dockets"] == 18
    # 8 distinct suppliers in seed POs + docket headers
    assert data["supplier_count"] >= 6


# ---------------------------------------------------------------------------
# Rate review (rate feedback loop) tests
# ---------------------------------------------------------------------------

def _summary_items(client, pid, supplier="Test Supplier"):
    resp = client.get(
        f"/api/projects/{pid}/docket-summary?supplier={supplier.replace(' ', '+')}"
    )
    data = json.loads(resp.data)
    items = []
    for g in data["groups"]:
        items.extend(g["items"])
    return items, data


def test_summary_rows_carry_line_ids_and_rate(client):
    pid, ids = _create_test_dockets(client)
    items, data = _summary_items(client, pid)
    assert len(items) > 0
    for item in items:
        assert "line_ids" in item
        assert "rate" in item
        assert "resource_id" in item
        # line_ids is a comma-joined string of ints
        assert all(p.isdigit() for p in str(item["line_ids"]).split(","))


def test_rerate_lines(client):
    pid, ids = _create_test_dockets(client)
    items, before = _summary_items(client, pid)
    target = items[0]
    line_ids = [int(x) for x in str(target["line_ids"]).split(",")]
    new_rate = round((target["rate"] or 0) + 10, 2)

    resp = client.post(
        f"/api/projects/{pid}/rerate",
        json={"line_ids": line_ids, "new_rate": new_rate},
        content_type="application/json",
    )
    assert resp.status_code == 200
    result = json.loads(resp.data)
    assert result["updated_lines"] == len(line_ids)
    assert result["old_rate"] == target["rate"]

    # Summary now reflects the new rate and re-valued subtotal
    items_after, after = _summary_items(client, pid)
    match = [i for i in items_after if str(i["line_ids"]) == str(target["line_ids"])
             or set(str(i["line_ids"]).split(",")) & set(map(str, line_ids))]
    assert any(abs(i["rate"] - new_rate) < 0.01 for i in match)
    assert after["grand_total"] != before["grand_total"]


def test_rerate_updates_standard_rate(client):
    pid, ids = _create_test_dockets(client)

    # Create a resource and a docket carrying it
    res = json.loads(client.post("/api/resources", json={
        "description": "Test Roller", "unit": "Hr", "standard_rate": 150,
    }, content_type="application/json").data)
    client.post(f"/api/projects/{pid}/dockets", json={
        "date": "2025-02-10", "supplier_name": "Test Supplier",
        "docket_number": "RGC-T009",
        "lines": [{"resource_id": res["id"], "description": "Test Roller",
                   "qty": 4, "unit": "Hr", "rate": 150}],
    }, content_type="application/json")

    items, _ = _summary_items(client, pid)
    target = next(i for i in items if i["resource_id"] == res["id"])
    line_ids = [int(x) for x in str(target["line_ids"]).split(",")]

    resp = client.post(f"/api/projects/{pid}/rerate", json={
        "line_ids": line_ids, "new_rate": 165,
        "resource_id": res["id"], "update_standard": True,
    }, content_type="application/json")
    assert resp.status_code == 200
    result = json.loads(resp.data)
    assert result["standard_updated"] is True
    assert result["old_standard_rate"] == 150

    updated = json.loads(client.get(f"/api/resources/{res['id']}").data)
    assert updated["standard_rate"] == 165


def test_rerate_add_resource_from_free_text(client):
    pid, ids = _create_test_dockets(client)

    # Free-text docket line (no resource), unrated
    client.post(f"/api/projects/{pid}/dockets", json={
        "date": "2025-02-11", "supplier_name": "Test Supplier",
        "docket_number": "RGC-T011",
        "lines": [{"description": "Rock breaker attachment",
                   "qty": 6, "unit": "Hr", "rate": 0}],
    }, content_type="application/json")

    items, _ = _summary_items(client, pid)
    target = next(i for i in items if i["resource_desc"] == "Rock breaker attachment")
    assert target["resource_id"] is None
    line_ids = [int(x) for x in str(target["line_ids"]).split(",")]

    resp = client.post(f"/api/projects/{pid}/rerate", json={
        "line_ids": line_ids, "new_rate": 95,
        "add_resource": {"description": "Rock breaker attachment",
                         "unit": "Hr", "supplier_name": "Test Supplier"},
    }, content_type="application/json")
    assert resp.status_code == 200
    result = json.loads(resp.data)
    assert result["new_resource_id"] is not None

    # Resource exists with the invoice rate as standard
    res = json.loads(client.get(f"/api/resources/{result['new_resource_id']}").data)
    assert res["standard_rate"] == 95
    assert res["description"] == "Rock breaker attachment"

    # Lines are now linked to the resource
    items_after, _ = _summary_items(client, pid)
    relinked = next(i for i in items_after
                    if i["resource_desc"] == "Rock breaker attachment")
    assert relinked["resource_id"] == result["new_resource_id"]
    assert abs(relinked["rate"] - 95) < 0.01


def test_rerate_validation(client):
    pid, ids = _create_test_dockets(client)

    # Missing line_ids
    resp = client.post(f"/api/projects/{pid}/rerate",
                       json={"new_rate": 100}, content_type="application/json")
    assert resp.status_code == 400

    # Negative rate
    resp = client.post(f"/api/projects/{pid}/rerate",
                       json={"line_ids": [1], "new_rate": -5},
                       content_type="application/json")
    assert resp.status_code == 400

    # Lines from another project (seed project 1)
    resp = client.post(f"/api/projects/{pid}/rerate",
                       json={"line_ids": [1], "new_rate": 100},
                       content_type="application/json")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Backup and Excel export tests
# ---------------------------------------------------------------------------

def test_backup_download(client):
    resp = client.get("/api/backup")
    assert resp.status_code == 200
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    # Valid SQLite file magic
    assert resp.data[:16] == b"SQLite format 3\x00"


def test_export_dockets_xlsx(client):
    import io as _io
    from openpyxl import load_workbook

    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/projects/{pid}/dockets/export-xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.content_type

    wb = load_workbook(_io.BytesIO(resp.data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Date" in headers
    assert "Amount" in headers
    assert ws.max_row == 5  # header + 4 line items


def test_resources_export_csv(client):
    resp = client.get("/api/resources/export-csv")
    assert resp.status_code == 200
    assert "text/csv" in resp.content_type
    lines = resp.data.decode("utf-8").strip().split("\n")
    assert "Item" in lines[0]
    assert len(lines) > 1  # seed resources present


def test_resources_export_xlsx(client):
    import io as _io
    from openpyxl import load_workbook

    resp = client.get("/api/resources/export-xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.content_type
    wb = load_workbook(_io.BytesIO(resp.data))
    headers = [c.value for c in wb.active[1]]
    assert "Standard Rate" in headers


def test_resources_import_csv(client):
    csv_text = (
        "Description,Unit,Supplier,Standard Rate,Category\n"
        "Imported Grader 14ft,Hr,Test Plant Co,210,Plant\n"
        "Imported Spotter,Hr,,72.50,Labour\n"
    )
    resp = client.post("/api/resources/import-csv",
                       json={"csv_text": csv_text},
                       content_type="application/json")
    assert resp.status_code == 200
    result = json.loads(resp.data)
    assert result["created"] == 2
    assert result["skipped"] == 0

    # Re-import skips both as duplicates
    resp = client.post("/api/resources/import-csv",
                       json={"csv_text": csv_text},
                       content_type="application/json")
    result = json.loads(resp.data)
    assert result["created"] == 0
    assert result["skipped"] == 2

    # Imported resource is queryable with the right rate
    resources = json.loads(client.get("/api/resources").data)
    grader = next(r for r in resources if r["description"] == "Imported Grader 14ft")
    assert grader["standard_rate"] == 210
    assert grader["supplier_name"] == "Test Plant Co"


def test_resource_details_roundtrip(client):
    """Item + Description split: details field saves, updates, and exports."""
    res = json.loads(client.post("/api/resources", json={
        "description": "Excavator 14T", "unit": "Hr",
        "details": "Cat 314, rubber tracked, long arm",
    }, content_type="application/json").data)
    assert res["details"] == "Cat 314, rubber tracked, long arm"

    updated = json.loads(client.put(f"/api/resources/{res['id']}", json={
        "details": "Cat 314, steel tracked",
    }, content_type="application/json").data)
    assert updated["details"] == "Cat 314, steel tracked"
    assert updated["description"] == "Excavator 14T"

    # Export carries both columns
    text = client.get("/api/resources/export-csv").data.decode("utf-8")
    header = text.strip().split("\n")[0]
    assert "Item" in header and "Description" in header
    assert "Cat 314, steel tracked" in text


def test_resources_import_item_description_split(client):
    """With an Item column, Description maps to the details field."""
    csv_text = (
        "Item,Description,Unit,Supplier,Standard Rate,Category\n"
        "Dozer D6,Cat D6 XE with slope assist,Hr,Test Plant Co,310,Plant\n"
    )
    resp = client.post("/api/resources/import-csv",
                       json={"csv_text": csv_text},
                       content_type="application/json")
    assert json.loads(resp.data)["created"] == 1
    resources = json.loads(client.get("/api/resources").data)
    dozer = next(r for r in resources if r["description"] == "Dozer D6")
    assert dozer["details"] == "Cat D6 XE with slope assist"


def test_resources_import_legacy_description_only(client):
    """Without an Item column, Description is the item name (old exports)."""
    csv_text = (
        "Description,Unit,Standard Rate\n"
        "Legacy Pump,Day,95\n"
    )
    resp = client.post("/api/resources/import-csv",
                       json={"csv_text": csv_text},
                       content_type="application/json")
    assert json.loads(resp.data)["created"] == 1
    resources = json.loads(client.get("/api/resources").data)
    pump = next(r for r in resources if r["description"] == "Legacy Pump")
    assert pump["details"] is None


def test_supplier_registered_from_po(client):
    """A new supplier typed on a PO joins the suppliers source-of-truth list."""
    before = json.loads(client.get("/api/projects/1/suppliers").data)
    assert "Zephyr Plant Hire" not in before

    client.post("/api/projects/1/purchase-orders", json={
        "number": "PO-NEW-1", "supplier_name": "Zephyr Plant Hire", "value": 1000,
    }, content_type="application/json")

    after = json.loads(client.get("/api/projects/1/suppliers").data)
    assert "Zephyr Plant Hire" in after


def test_supplier_name_canonicalised(client):
    """Typing a different-case variant resolves to the stored canonical name."""
    # 'Ironbark Quarries' is in the seed; enter a lowercase variant on a PO
    po = json.loads(client.post("/api/projects/1/purchase-orders", json={
        "number": "PO-NEW-2", "supplier_name": "ironbark quarries", "value": 500,
    }, content_type="application/json").data)
    assert po["supplier_name"] == "Ironbark Quarries"  # canonical spelling, not the typed case

    # And no duplicate variant was added to the list
    suppliers = json.loads(client.get("/api/projects/1/suppliers").data)
    assert suppliers.count("Ironbark Quarries") == 1
    assert "ironbark quarries" not in suppliers


def test_resource_supplier_registered(client):
    res = json.loads(client.post("/api/resources", json={
        "description": "Crane 50T", "unit": "Day", "supplier_name": "Skyhook Cranes",
    }, content_type="application/json").data)
    assert res["supplier_name"] == "Skyhook Cranes"
    suppliers = json.loads(client.get("/api/projects/1/suppliers").data)
    assert "Skyhook Cranes" in suppliers


def test_resources_import_requires_columns(client):
    resp = client.post("/api/resources/import-csv",
                       json={"csv_text": "Name,Price\nThing,5\n"},
                       content_type="application/json")
    assert resp.status_code == 400


def test_docket_summary_xlsx(client):
    import io as _io
    from openpyxl import load_workbook

    pid, ids = _create_test_dockets(client)
    resp = client.get(
        f"/api/projects/{pid}/docket-summary/xlsx?supplier=Test+Supplier"
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.content_type

    wb = load_workbook(_io.BytesIO(resp.data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Rate" in headers
    assert "Subtotal" in headers
    # Last row is the grand total
    last = [c.value for c in ws[ws.max_row]]
    assert "Grand Total" in last


# ---------------------------------------------------------------------------
# Scan serving / source_filepath tests
# ---------------------------------------------------------------------------


def test_source_filepath_persists(client):
    """source_filepath is stored on create and returned on GET."""
    pid, _ = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets",
        json={
            "date": "2025-06-01",
            "docket_number": "FP-001",
            "source_hash": "aabbccdd",
            "source_filename": "scan.pdf",
            "source_filepath": "C:/Scans/June/scan.pdf",
            "lines": [{"description": "Item", "qty": 1, "rate": 50}],
        },
        content_type="application/json",
    )
    assert resp.status_code == 201
    did = json.loads(resp.data)["id"]

    docket = json.loads(client.get(f"/api/dockets/{did}").data)
    assert docket["source_filepath"] == "C:/Scans/June/scan.pdf"
    assert docket["source_hash"] == "aabbccdd"
    assert docket["source_filename"] == "scan.pdf"


def test_scan_endpoint_404_no_filepath(client):
    """GET /api/scans/<id> returns 404 when no filepath is stored."""
    pid, ids = _create_test_dockets(client)
    resp = client.get(f"/api/scans/{ids[0]}")
    assert resp.status_code == 404
    data = json.loads(resp.data)
    assert "No scan" in data["error"]


def test_scan_endpoint_404_missing_file(client):
    """GET /api/scans/<id> returns 404 when file doesn't exist on disk."""
    pid, _ = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets",
        json={
            "date": "2025-06-01",
            "source_hash": "deadbeef",
            "source_filepath": "/nonexistent/scan.pdf",
            "lines": [{"description": "X", "qty": 1, "rate": 10}],
        },
        content_type="application/json",
    )
    did = json.loads(resp.data)["id"]
    scan_resp = client.get(f"/api/scans/{did}")
    assert scan_resp.status_code == 404
    assert "not found" in json.loads(scan_resp.data)["error"].lower()


def test_scan_endpoint_serves_file(client):
    """GET /api/scans/<id> serves the file when path and hash match."""
    pid, _ = _create_test_dockets(client)

    import hashlib
    content = b"fake PDF content for testing"
    h = hashlib.sha256(content).hexdigest()

    fd, path = tempfile.mkstemp(suffix=".pdf")
    os.write(fd, content)
    os.close(fd)
    try:
        resp = client.post(
            f"/api/projects/{pid}/dockets",
            json={
                "date": "2025-06-01",
                "source_hash": h,
                "source_filepath": path,
                "lines": [{"description": "Y", "qty": 1, "rate": 10}],
            },
            content_type="application/json",
        )
        did = json.loads(resp.data)["id"]
        scan_resp = client.get(f"/api/scans/{did}")
        assert scan_resp.status_code == 200
        assert scan_resp.data == content
        scan_resp.close()
    finally:
        try:
            os.unlink(path)
        except PermissionError:
            pass


def test_scan_endpoint_hash_mismatch(client):
    """GET /api/scans/<id> returns 409 when the file hash changed."""
    pid, _ = _create_test_dockets(client)

    fd, path = tempfile.mkstemp(suffix=".pdf")
    os.write(fd, b"original")
    os.close(fd)
    try:
        resp = client.post(
            f"/api/projects/{pid}/dockets",
            json={
                "date": "2025-06-01",
                "source_hash": "wrong_hash",
                "source_filepath": path,
                "lines": [{"description": "Z", "qty": 1, "rate": 10}],
            },
            content_type="application/json",
        )
        did = json.loads(resp.data)["id"]
        scan_resp = client.get(f"/api/scans/{did}")
        assert scan_resp.status_code == 409
        data = json.loads(scan_resp.data)
        assert "mismatch" in data["error"].lower()
        scan_resp.close()
    finally:
        try:
            os.unlink(path)
        except PermissionError:
            pass


def test_unassign_scan(client):
    """PUT with null source fields clears the association."""
    pid, _ = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets",
        json={
            "date": "2025-06-01",
            "docket_number": "UNSCAN-001",
            "source_hash": "aabb",
            "source_filename": "doc.pdf",
            "source_filepath": "/path/doc.pdf",
            "lines": [{"description": "A", "qty": 1, "rate": 10}],
        },
        content_type="application/json",
    )
    did = json.loads(resp.data)["id"]

    # Unassign
    client.put(
        f"/api/dockets/{did}",
        json={
            "source_hash": None,
            "source_filename": None,
            "source_filepath": None,
        },
        content_type="application/json",
    )

    docket = json.loads(client.get(f"/api/dockets/{did}").data)
    assert docket["source_hash"] is None
    assert docket["source_filename"] is None
    assert docket["source_filepath"] is None


def test_check_path(client):
    """check-path tells the UI whether a pasted path resolves to a real file."""
    # A real file on disk → absolute + exists
    fd, path = tempfile.mkstemp(suffix=".pdf")
    os.write(fd, b"data")
    os.close(fd)
    try:
        resp = client.post("/api/check-path", json={"path": path})
        data = json.loads(resp.data)
        assert data["absolute"] is True
        assert data["exists"] is True
    finally:
        os.unlink(path)

    # A bare relative folder name (the trap) → not absolute, not found
    resp = client.post("/api/check-path", json={"path": "Corpus/scan.pdf"})
    data = json.loads(resp.data)
    assert data["absolute"] is False
    assert data["exists"] is False

    # An absolute path to a missing file → absolute but not found
    resp = client.post(
        "/api/check-path", json={"path": "/nonexistent/dir/scan.pdf"}
    )
    data = json.loads(resp.data)
    assert data["exists"] is False

    # Empty / missing path → both false, no crash
    resp = client.post("/api/check-path", json={})
    data = json.loads(resp.data)
    assert data["exists"] is False
    assert data["absolute"] is False


def test_list_folder(client):
    """list-folder fingerprints scannable files at a real path, with absolute
    paths, and ignores non-scan files."""
    import hashlib
    d = tempfile.mkdtemp()
    try:
        with open(os.path.join(d, "a.pdf"), "wb") as f:
            f.write(b"pdf-a")
        with open(os.path.join(d, "b.png"), "wb") as f:
            f.write(b"png-b")
        with open(os.path.join(d, "notes.txt"), "wb") as f:
            f.write(b"ignore me")

        resp = client.post("/api/list-folder", json={"path": d})
        assert resp.status_code == 200
        data = json.loads(resp.data)
        names = sorted(f["name"] for f in data["files"])
        assert names == ["a.pdf", "b.png"]  # .txt skipped
        for f in data["files"]:
            assert os.path.isabs(f["path"])
            assert len(f["hash"]) == 64
        # Hash matches a direct computation
        amatch = next(f for f in data["files"] if f["name"] == "a.pdf")
        assert amatch["hash"] == hashlib.sha256(b"pdf-a").hexdigest()
    finally:
        import shutil
        shutil.rmtree(d, ignore_errors=True)


def test_list_folder_missing(client):
    """list-folder 404s on a folder that doesn't exist."""
    resp = client.post("/api/list-folder", json={"path": "/no/such/folder/xyz"})
    assert resp.status_code == 404


def test_scan_file_serves_browsed(client):
    """scan-file serves a file once its folder has been browsed."""
    d = tempfile.mkdtemp()
    try:
        p = os.path.join(d, "doc.pdf")
        with open(p, "wb") as f:
            f.write(b"hello scan")
        # Browsing the folder authorises previews from it
        client.post("/api/list-folder", json={"path": d})
        resp = client.get("/api/scan-file", query_string={"path": p})
        assert resp.status_code == 200
        assert resp.data == b"hello scan"
        resp.close()
    finally:
        import shutil
        shutil.rmtree(d, ignore_errors=True)


def test_scan_file_rejects_unbrowsed(client):
    """scan-file refuses a path whose folder hasn't been browsed (403)."""
    d = tempfile.mkdtemp()
    try:
        p = os.path.join(d, "secret.pdf")
        with open(p, "wb") as f:
            f.write(b"x")
        # No list-folder call for this dir → not authorised
        resp = client.get("/api/scan-file", query_string={"path": p})
        assert resp.status_code == 403
    finally:
        import shutil
        shutil.rmtree(d, ignore_errors=True)


def test_pick_folder_subprocess(client, monkeypatch):
    """pick-folder shells out to a fresh process; the test hook returns a real
    folder so the full subprocess + listing path runs without a GUI."""
    import shutil
    d = tempfile.mkdtemp()
    try:
        with open(os.path.join(d, "scan.pdf"), "wb") as f:
            f.write(b"x")
        monkeypatch.setenv("DCT_PICK_FOLDER_TEST", d)
        resp = client.post("/api/pick-folder", json={})
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data["available"] is True
        assert data.get("cancelled") is not True
        assert os.path.samefile(data["folder"], d)
        assert any(f["name"] == "scan.pdf" for f in data["files"])
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_pick_folder_cancelled(client, monkeypatch):
    """An empty pick (user cancelled the dialog) → cancelled, no files."""
    monkeypatch.setenv("DCT_PICK_FOLDER_TEST", "__CANCEL__")
    resp = client.post("/api/pick-folder", json={})
    data = json.loads(resp.data)
    assert data["available"] is True
    assert data["cancelled"] is True


def test_fs_browse_disabled_gate(client, monkeypatch):
    """With DCT_OS_NO_FS_BROWSE set (e.g. the public demo), the filesystem
    browse endpoints refuse to expose the server's files."""
    monkeypatch.setenv("DCT_OS_NO_FS_BROWSE", "1")

    resp = client.post("/api/pick-folder", json={})
    assert json.loads(resp.data)["available"] is False

    resp = client.post("/api/list-folder", json={"path": "/"})
    assert resp.status_code == 403

    resp = client.get("/api/scan-file", query_string={"path": "/etc/hosts"})
    assert resp.status_code == 403


def test_scan_root_restriction(client, monkeypatch):
    """With DCT_OS_SCAN_ROOT set, /api/scans refuses paths outside the root
    (the demo hardening) but still serves ones inside it."""
    import shutil
    root = tempfile.mkdtemp()
    outside = tempfile.mkdtemp()
    try:
        inside_path = os.path.join(root, "ok.pdf")
        with open(inside_path, "wb") as f:
            f.write(b"inside")
        outside_path = os.path.join(outside, "secret.pdf")
        with open(outside_path, "wb") as f:
            f.write(b"secret")

        pid, _ = _create_test_dockets(client)
        # Docket pointing inside the root, and one pointing outside it.
        d_in = json.loads(client.post(f"/api/projects/{pid}/dockets", json={
            "date": "2025-06-01", "source_filepath": inside_path,
            "lines": [{"description": "x", "qty": 1, "rate": 1}]}).data)["id"]
        d_out = json.loads(client.post(f"/api/projects/{pid}/dockets", json={
            "date": "2025-06-01", "source_filepath": outside_path,
            "lines": [{"description": "y", "qty": 1, "rate": 1}]}).data)["id"]

        monkeypatch.setenv("DCT_OS_SCAN_ROOT", root)
        r_in = client.get(f"/api/scans/{d_in}")
        assert r_in.status_code == 200
        assert r_in.data == b"inside"
        r_in.close()
        r_out = client.get(f"/api/scans/{d_out}")
        assert r_out.status_code == 403  # outside the root → refused
    finally:
        shutil.rmtree(root, ignore_errors=True)
        shutil.rmtree(outside, ignore_errors=True)


def test_scan_endpoint_404_after_unassign(client):
    """After unassigning, the scan endpoint returns 404."""
    pid, _ = _create_test_dockets(client)
    resp = client.post(
        f"/api/projects/{pid}/dockets",
        json={
            "date": "2025-06-01",
            "source_hash": "cc",
            "source_filepath": "/some/file.pdf",
            "lines": [{"description": "B", "qty": 1, "rate": 10}],
        },
        content_type="application/json",
    )
    did = json.loads(resp.data)["id"]

    # Unassign
    client.put(
        f"/api/dockets/{did}",
        json={"source_hash": None, "source_filepath": None},
        content_type="application/json",
    )

    scan_resp = client.get(f"/api/scans/{did}")
    assert scan_resp.status_code == 404

